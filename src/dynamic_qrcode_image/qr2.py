# 生成带logo的二维码图片
from openpyxl import load_workbook
from time import strftime,localtime,time
from .qrc import QRC        
import config

# 根据配置生成文件名
def generate_uri(row,s='',char='-'):
    for k in config.name:
        n=ord(config.name[k])-ord("A")
        s+=f'{row[n].value}{char}'
    return s[:-1]

# 根据配置生成地址串
def generate_str(row):
    s=config.baseurl
    for k in config.column:
        n=ord(config.column[k])-ord("A")
        s+=f'{k}={row[n].value}&'
    return s[:-1]

from sys import argv
if __name__ == '__main__': 
    wb = load_workbook(filename = f'{argv[1]}.xlsx')
    # 默认第一张表
    sheet_ranges = wb[wb.sheetnames[0]]
    qrc = QRC(zip_name=f"{config.zip}-{strftime('%H%M%S',localtime(time()))}")
    for row in sheet_ranges.rows:
        qrc.zip(generate_uri(row),content=generate_str(row))
    if len(argv)>2:
        print(argv[2],end='')
