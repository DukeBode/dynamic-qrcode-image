# 生成带logo的二维码图片
from os import getcwd,listdir,remove
from openpyxl import load_workbook
import config
from time import strftime,localtime,time
from zipfile import ZipFile
from qrc import QRC

# 生成二维码
def zip_qrcode(str,uri=False,zip=False,file='png'):
    img = QRC().getCode(str,uri)
    name=f'{uri}.{file}'
    img.save(name)
    if zip:
        with ZipFile(zip, 'a') as myzip:
            myzip.write(name)
            remove(name)
        

# 根据配置生成文件名
def generate_uri(row,s='',char='-'):
    for k in config.name:
        n=ord(config.name[k])-ord("A")
        s+=f'{row[n].value}{char}'
    return s[:-1]

# 根据配置生成地址串
def generate_str(row):
    s=config.baseurl
    for k in config.column:
        n=ord(config.column[k])-ord("A")
        s+=f'{k}={row[n].value}&'
    return s[:-1]

from sys import argv
if __name__ == '__main__': 
    wb = load_workbook(filename = f'{argv[1]}.xlsx')
    # 默认第一张表
    sheet_ranges = wb[wb.sheetnames[0]]
    zip_name=f"{config.zip}-{strftime('%H%M%S',localtime(time()))}.zip"
    for row in sheet_ranges.rows:
        zip_qrcode(generate_str(row),generate_uri(row),zip_name)
    if len(argv)>2:
        print(argv[2],end='')
    print(zip_name)

